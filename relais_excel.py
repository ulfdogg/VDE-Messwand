"""
VDE Messwand - Excel Import/Export für Relais-Konfiguration
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from relais_manager import get_all_relais_config, bulk_update_relais


def create_excel_template(include_current_config=False):
    """
    Erstellt eine Excel-Vorlage für Relais-Konfiguration

    Args:
        include_current_config: Wenn True, aktuelle Konfiguration einschließen

    Returns:
        BytesIO Objekt mit Excel-Datei
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relais-Konfiguration"

    # Header-Style
    header_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    headers = ["Relais-Nr", "Gruppen-Nr", "Name", "Kategorie", "Stromkreis"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Spaltenbreiten
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Daten
    if include_current_config:
        # Aktuelle Konfiguration laden
        config = get_all_relais_config()
        for relay_num in range(64):
            relay_data = config.get(relay_num, {})
            row = relay_num + 2

            ws.cell(row=row, column=1, value=relay_num).border = thin_border
            ws.cell(row=row, column=2, value=relay_data.get('group_number', 0)).border = thin_border
            ws.cell(row=row, column=3, value=relay_data.get('name', '')).border = thin_border
            ws.cell(row=row, column=4, value=relay_data.get('category', '')).border = thin_border
            ws.cell(row=row, column=5, value=relay_data.get('stromkreis', '')).border = thin_border
    else:
        # Leere Vorlage mit allen 64 Relais
        for relay_num in range(64):
            row = relay_num + 2
            ws.cell(row=row, column=1, value=relay_num).border = thin_border
            ws.cell(row=row, column=2, value=0).border = thin_border
            ws.cell(row=row, column=3, value='').border = thin_border
            ws.cell(row=row, column=4, value='').border = thin_border
            ws.cell(row=row, column=5, value='').border = thin_border

    # Hinweise auf extra Sheet
    ws_info = wb.create_sheet("Hinweise")
    ws_info.column_dimensions['A'].width = 50

    info_text = [
        ("VDE Messwand - Relais-Konfiguration", True),
        ("", False),
        ("Anleitung:", True),
        ("1. Füllen Sie die Tabelle 'Relais-Konfiguration' aus", False),
        ("2. Relais-Nr: 0-63 (nicht ändern!)", False),
        ("3. Gruppen-Nr: 0 = einzeln, 1-20 = Gruppennummer", False),
        ("4. Name: Beliebige Beschreibung (z.B. 'CEE L1')", False),
        ("5. Kategorie: z.B. RISO, Zi, Zs, RCD", False),
        ("6. Stromkreis: z.B. L1, L2, L3, N, PE", False),
        ("", False),
        ("Beispiele:", True),
        ("Relais 0-2 gruppieren: Alle mit Gruppen-Nr = 1", False),
        ("CEE 3-Phasen: L1, L2, L3 in Gruppe 1, N und PE einzeln", False),
        ("", False),
        ("Hinweis:", True),
        ("- Relais mit gleicher Gruppen-Nr werden zusammen geschaltet", False),
        ("- Gruppe 0 bedeutet 'einzelnes Relais'", False),
        ("- Leere Zeilen werden ignoriert", False),
    ]

    for row_num, (text, is_bold) in enumerate(info_text, 1):
        cell = ws_info.cell(row=row_num, column=1, value=text)
        if is_bold:
            cell.font = Font(bold=True, size=12, color="FF4444")
        cell.alignment = Alignment(wrap_text=True)

    # Speichere in BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def import_from_excel(file_stream):
    """
    Importiert Relais-Konfiguration aus Excel-Datei

    Args:
        file_stream: File-Stream der hochgeladenen Excel-Datei

    Returns:
        Dictionary mit success, message und optional imported_count
    """
    try:
        wb = load_workbook(file_stream)
        ws = wb.active

        # Überprüfe Header
        expected_headers = ["Relais-Nr", "Gruppen-Nr", "Name", "Kategorie", "Stromkreis"]
        actual_headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]

        if actual_headers != expected_headers:
            # Detaillierte Fehlermeldung
            header_info = '\n'.join([f'Spalte {i+1}: Erwartet "{expected_headers[i]}", gefunden "{actual_headers[i]}"'
                                     for i in range(len(expected_headers))
                                     if i < len(actual_headers) and actual_headers[i] != expected_headers[i]])
            return {
                'success': False,
                'message': f'Ungültige Excel-Datei Header:\n{header_info}\n\nBitte verwenden Sie die richtige Vorlage!'
            }

        # Lese Daten
        updates = {}
        errors = []

        for row_num in range(2, ws.max_row + 1):
            relay_num = ws.cell(row=row_num, column=1).value
            group_number = ws.cell(row=row_num, column=2).value
            name = ws.cell(row=row_num, column=3).value
            category = ws.cell(row=row_num, column=4).value
            stromkreis = ws.cell(row=row_num, column=5).value

            # Validierung
            if relay_num is None:
                continue

            try:
                relay_num = int(relay_num)
                group_number = int(group_number) if group_number is not None else 0

                if not (0 <= relay_num <= 63):
                    errors.append(f"Zeile {row_num}: Ungültige Relais-Nr {relay_num}")
                    continue

                if not (0 <= group_number <= 20):
                    errors.append(f"Zeile {row_num}: Ungültige Gruppen-Nr {group_number}")
                    continue

                # Konvertiere Werte zu String, handle None-Werte korrekt
                name_str = '' if name is None else str(name).strip()
                category_str = '' if category is None else str(category).strip()
                stromkreis_str = '' if stromkreis is None else str(stromkreis).strip()

                updates[relay_num] = {
                    'group_number': group_number,
                    'name': name_str,
                    'category': category_str,
                    'stromkreis': stromkreis_str
                }


            except (ValueError, TypeError) as e:
                errors.append(f"Zeile {row_num}: Fehler beim Lesen der Daten - {str(e)}")
                continue

        if errors:
            error_count = len(errors)
            error_preview = '\n'.join(errors[:10])  # Max 10 Fehler anzeigen
            more_errors = f'\n... und {error_count - 10} weitere Fehler' if error_count > 10 else ''
            return {
                'success': False,
                'message': f'Fehler beim Importieren ({error_count} Fehler):\n{error_preview}{more_errors}'
            }

        if not updates:
            return {
                'success': False,
                'message': 'Keine gültigen Daten in der Excel-Datei gefunden. Überprüfen Sie, ob die Datei Daten enthält und das richtige Format hat.'
            }

        # Importiere Daten
        result = bulk_update_relais(updates)

        if result['success']:
            return {
                'success': True,
                'message': f'Erfolgreich {len(updates)} Relais importiert',
                'imported_count': len(updates)
            }
        else:
            return {
                'success': False,
                'message': result.get('message', 'Fehler beim Speichern der Daten')
            }

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Excel Import Error: {error_details}")  # Für Debug-Log
        return {
            'success': False,
            'message': f'Fehler beim Verarbeiten der Excel-Datei:\n{str(e)}\n\nStellen Sie sicher, dass:\n- Die Datei im Excel-Format (.xlsx oder .xls) ist\n- Die Datei die richtige Vorlage verwendet\n- Die Datei nicht beschädigt ist'
        }


def get_predefined_templates():
    """
    Gibt vordefinierte Excel-Vorlagen zurück (bereits konfiguriert)

    Returns:
        Liste von Template-Dictionaries mit Namen und Konfigurationen
    """
    return [
        {
            'id': 'cee_3phase',
            'name': 'CEE 3-Phasen Steckdose',
            'description': 'Standard CEE 16A Steckdose mit 3 Phasen gruppiert',
            'config': {
                0: {'group_number': 1, 'name': 'CEE L1', 'category': 'RISO', 'stromkreis': 'L1'},
                1: {'group_number': 1, 'name': 'CEE L2', 'category': 'RISO', 'stromkreis': 'L2'},
                2: {'group_number': 1, 'name': 'CEE L3', 'category': 'RISO', 'stromkreis': 'L3'},
                3: {'group_number': 0, 'name': 'CEE N', 'category': 'Zi', 'stromkreis': 'N'},
                4: {'group_number': 0, 'name': 'CEE PE', 'category': 'Zi', 'stromkreis': 'PE'}
            }
        },
        {
            'id': 'herd_4wire',
            'name': 'Herdanschluss 4-adrig',
            'description': 'Herdanschlussdose mit L1, L2, L3, N gruppiert',
            'config': {
                10: {'group_number': 2, 'name': 'Herd L1', 'category': 'RISO', 'stromkreis': 'L1'},
                11: {'group_number': 2, 'name': 'Herd L2', 'category': 'RISO', 'stromkreis': 'L2'},
                12: {'group_number': 2, 'name': 'Herd L3', 'category': 'RISO', 'stromkreis': 'L3'},
                13: {'group_number': 2, 'name': 'Herd N', 'category': 'Zi', 'stromkreis': 'N'}
            }
        },
        {
            'id': 'wallbox_3phase',
            'name': 'Wallbox 3-Phasen',
            'description': 'E-Auto Ladestation mit 3 Phasen und PE',
            'config': {
                50: {'group_number': 3, 'name': 'Wallbox L1', 'category': 'RISO', 'stromkreis': 'L1'},
                51: {'group_number': 3, 'name': 'Wallbox L2', 'category': 'RISO', 'stromkreis': 'L2'},
                52: {'group_number': 3, 'name': 'Wallbox L3', 'category': 'RISO', 'stromkreis': 'L3'},
                53: {'group_number': 0, 'name': 'Wallbox PE', 'category': 'Zi', 'stromkreis': 'PE'}
            }
        }
    ]


def create_predefined_template_excel(template_id):
    """
    Erstellt Excel-Datei mit vordefinierter Vorlage

    Args:
        template_id: ID der Vorlage

    Returns:
        BytesIO Objekt mit Excel-Datei oder None
    """
    templates = get_predefined_templates()

    for template in templates:
        if template['id'] == template_id:
            wb = Workbook()
            ws = wb.active
            ws.title = "Relais-Konfiguration"

            # Header-Style (wie oben)
            header_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Header
            headers = ["Relais-Nr", "Gruppen-Nr", "Name", "Kategorie", "Stromkreis"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Spaltenbreiten
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # Füge alle 64 Relais hinzu, befülle nur die aus der Vorlage
            config = template['config']
            for relay_num in range(64):
                row = relay_num + 2

                if relay_num in config:
                    relay_data = config[relay_num]
                    ws.cell(row=row, column=1, value=relay_num).border = thin_border
                    ws.cell(row=row, column=2, value=relay_data['group_number']).border = thin_border
                    ws.cell(row=row, column=3, value=relay_data['name']).border = thin_border
                    ws.cell(row=row, column=4, value=relay_data['category']).border = thin_border
                    ws.cell(row=row, column=5, value=relay_data['stromkreis']).border = thin_border
                else:
                    ws.cell(row=row, column=1, value=relay_num).border = thin_border
                    ws.cell(row=row, column=2, value=0).border = thin_border
                    ws.cell(row=row, column=3, value='').border = thin_border
                    ws.cell(row=row, column=4, value='').border = thin_border
                    ws.cell(row=row, column=5, value='').border = thin_border

            # Speichere in BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            return excel_file

    return None
